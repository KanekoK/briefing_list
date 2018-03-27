import urllib.request
import openpyxl as px
from bs4 import BeautifulSoup

# URLの設定
url = "https://job.rikunabi.com/2019/search/seminar/result/?dd=20180406&k=11&k=12&k=13&k=14&pn="
# Sheet名の設定
sheet_name = "4月6日"

# 媒体の設定
media = "リクナビ"



def access(url, cnt):
	html = urllib.request.urlopen(url+str(cnt))
	soup = BeautifulSoup(html, "html.parser")
	if soup.select("._vacantSeat ._vacantSeat-data-day"):
		return soup
	else:
		return False

def get_info_list(soup):
	event_list   = []
	area_list    = []
	company_list = []
	link_list    = []
	event_date = soup.select("._vacantSeat ._vacantSeat-data-day")
	area = soup.select("._vacantSeat ._vacantSeat-data-place")
	company_name = soup.select(".search-cassette .search-cassette-title a")
	link = soup.select(".search-cassette-footer .search-cassette-actionBar-cell_04 a.mod-btn")

	for s in event_date:
		event_list.append(s.string)
	for s in area:
		area_list.append(s.string.replace(" ", "").replace("\n", ""))
	for s in company_name:
		company_list.append(s.string)
	for s in link:
		link_list.append("https://job.rikunabi.com" + s.get("href"))
	return event_list, area_list, company_list, link_list

def insert_info_list(soup, event_list, area_list, company_list, link_list):
	event_date = soup.select("._vacantSeat ._vacantSeat-data-day")
	area = soup.select("._vacantSeat ._vacantSeat-data-place")
	company_name = soup.select(".search-cassette .search-cassette-title a")
	link = soup.select(".search-cassette-footer .search-cassette-actionBar-cell_04 a.mod-btn")

	for s in event_date:
		event_list.append(s.string)
	for s in area:
		area_list.append(s.string.replace(" ", "").replace("\n", ""))
	for s in company_name:
		company_list.append(s.string)
	for s in link:
		link_list.append("https://job.rikunabi.com" + s.get("href"))
	return event_list, area_list, company_list, link_list

def make_assoc(event_list, area_list, company_list, link_list):
	info = {}
	sheet, book = excel_sheet("説明会一覧.xlsx", "Sheet1")
	header_setting(sheet)
	for i in range(len(event_list)):
		sheet.cell(row=i+2, column=1, value="リクナビ")
		sheet.cell(row=i+2, column=2, value=event_list[i])
		sheet.cell(row=i+2, column=3, value=area_list[i])
		sheet.cell(row=i+2, column=4, value=company_list[i])
		sheet.cell(row=i+2, column=5, value=link_list[i])
		sheet.cell(row=i+2, column=5, value=link_list[i]).hyperlink = link_list[i]
	book.save('説明会一覧.xlsx')

def header_setting(sheet):
	sheet['A1'].value = '媒体'
	sheet['B1'].value = '日付'
	sheet['C1'].value = 'エリア'
	sheet['D1'].value = '会社名'
	sheet['E1'].value = '説明会リンク'

def excel_sheet(filepath, sheetname):
	book = px.load_workbook(filepath)
	sheet = book.create_sheet(title=sheet_name)
	# sheet = book.active
	return sheet, book

if __name__ == '__main__':
	flag = True
	cnt = 1
	soup = access(url, cnt)
	event_list, area_list, company_list, link_list = get_info_list(soup)

	while flag:
		cnt += 1
		soup = access(url, cnt)
		if soup == False:
			flag = False
		else:
			event_list, area_list, company_list, link_list = insert_info_list(soup, event_list, area_list, company_list, link_list)

	make_assoc(event_list, area_list, company_list, link_list)
	print("完了！")

